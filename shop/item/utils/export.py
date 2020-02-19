from django.http import HttpResponse
from django.core import serializers

from box.shop.item.api.serializers import ItemSerializer

from openpyxl import Workbook
from datetime import timedelta, datetime
import csv


count = 0


def export_items_to_xml():
  # https://github.com/vinitkumar/json2xml
  # https://github.com/quandyfactory/dicttoxml/tree/master/dist
  # !!!!
  # https://github.com/quandyfactory/dicttoxml
  # dicttoxml + rest_framework.serializers.ModelSerializer
  # https://stackoverflow.com/questions/36021526/converting-an-array-dict-to-xml-in-python
  # !!!!
  items = Item.objects.all()
  result = "<objects>"
  for item in items:
    result += form_item(item)
  result += "</objects>"
  # result  = serializers.serialize('xml', items)
  with open('items.xml', 'w') as f:
    f.write(result)
  return True


def form_item(item):
  item = f"""
  <object>
    <field name="meta_title" descr="{item._meta.get_field('meta_title').verbose_name}">
      {item.meta_title}
    </field>
    <field name="meta_descr" descr="{item._meta.get_field('meta_descr').verbose_name}">
      {item.meta_descr}
    </field>
    <field name="meta_key" descr="{item._meta.get_field('meta_key').verbose_name}">
      {item.meta_key}
    </field>
    <field name="title" descr="{item._meta.get_field('title').verbose_name}">
      {item.title}
    </field>
    <field name="description" descr="{item._meta.get_field('description').verbose_name}">
      {item.description}
    </field>
    <field name="code" descr="{item._meta.get_field('code').verbose_name}">
      {item.code}
    </field>
    <field name="slug" descr="{item._meta.get_field('slug').verbose_name}">
      {item.slug}
    </field>
    <field name="thumbnail" descr="{item._meta.get_field('thumbnail').verbose_name}">
      {item.thumbnail}
    </field>
    <field name="old_price" descr="{item._meta.get_field('old_price').verbose_name}">
      {item.old_price}
    </field>
    <field name="new_price" descr="{item._meta.get_field('new_price').verbose_name}">
      {item.new_price}
    </field>
    <field name="currency" descr="{item._meta.get_field('currency').verbose_name}">
      {item.currency}
    </field>
    <field name="category" descr="{item._meta.get_field('category').verbose_name}">
      {item.category}
    </field>
    <field name="in_stock" descr="{item._meta.get_field('in_stock').verbose_name}">
      {item.in_stock}
    </field>
    <field name="is_new" descr="{item._meta.get_field('is_new').verbose_name}">
      {item.is_new}
    </field>
    <field name="is_active" descr="{item._meta.get_field('is_active').verbose_name}">
      {item.is_active}
    </field>
    <field name="created" descr="{item._meta.get_field('created').verbose_name}">
      {item.created}
    </field>
    <field name="updated" descr="{item._meta.get_field('updated').verbose_name}">
      {item.updated}
    </field>
    <field name="order" descr="{item._meta.get_field('order').verbose_name}">
      {item.order}
    </field>
  </object>
  """
  return item 



class ExportMixin:

    def export_items_to_xlsx(self, request, queryset):
        from box.shop.item.models import Item, ItemCategory
        items = Item.objects.all()
        categories = ItemCategory.objects.all()
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f"attachment; filename={datetime.now().strftime('%Y-%m-%d')}.xlsx"


        
        workbook = Workbook()
        workbook.remove(workbook.active)




        worksheet1 = workbook.create_sheet(
            title='Товары',
            index=1,
        )
        columns = [
            "Мета_Заголовок",
            "Мета_Описание",
            "Мета_Ключевые_Слова",
            "Заголовок",
            "Описание",
            "Артикул",
            "Категория",
            "Ссылка",
            "Старая_Цена",
            "Новая_Цена",
            "Изображения",
        ]
        biggest_item = items.first()
        for item in items:
            if item.features.all().count() > biggest_item.features.all().count():
                biggest_item = item 
        for i in range(int(biggest_item.features.all().count())):
            columns.append("Название_Характеристики")
            columns.append("Значение_Характеристики")
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet1.cell(row=row_num, column=col_num)
            cell.value = column_title
        for item in items:
            row_num += 1
            row = [
                item.meta_title,
                item.meta_descr,
                item.meta_key,
                item.title,
                item.description,
                item.code,
                item.category.slug,
                item.slug,
                item.old_price,
                item.new_price,
                ','.join([image.image.url for image in item.images.all()]),
            ]
            features = []
            for feature in item.features.all():
                features.append(feature.name)
                features.append(feature.value)
            row.extend(features)
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet1.cell(row=row_num, column=col_num)
                cell.value = cell_value




        worksheet2 = workbook.create_sheet(
            title='Категории',
            index=2,
        )
        columns = [
            "title",
            "slug",
            "parent",
        ]
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = column_title
        for category in categories:
            row_num += 1
            parent_slug = ''
            if category.parent:
                parent_slug = category.parent.slug
            row = [
                category.title,
                category.slug,
                parent_slug,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet2.cell(row=row_num, column=col_num)
                cell.value = cell_value







        workbook.save(response)
        return response

    def export_items(self, request, queryset):
        meta = self.model._meta
        field_names = ['Места'] + [field.name for field in meta.fields]
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = f'attachement; filename={meta}.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in queryset:
            # seats = ','.join([seat_in_order.seat.number for seat_in_order in SeatInOrder.objects.filter(order=obj)])
            row = writer.writerow([getattr(obj, field) for field in field_names[1:]])
            # writer.writerow([seats] + [getattr(obj, field) for field in field_names[1:]])
            pass 
        return response
    
    def export_items_photoes(self, request, queryset):
      # https://thispointer.com/python-how-to-create-a-zip-archive-from-multiple-files-or-directory/
      import os 
      from zipfile import ZipFile, ZIP_DEFLATED
      from wsgiref.util import FileWrapper
      from django.conf import settings 

      images = []
      for item in queryset:
        images.extend(item.images.all() )
      print(len(images))

      with ZipFile('export.zip', 'w', ZIP_DEFLATED) as export_zip:
        for image in images:
          try:
            filename = settings.MEDIA_ROOT + image.image.url[6:]
            arcname = os.path.join('shop', 'item', image.item.slug, filename.split('/')[-1])
            export_zip.write(filename, arcname)
          except FileNotFoundError as e:
            print(e)
      response = HttpResponse(FileWrapper(open('export.zip', 'rb')), content_type='application/zip')
      response['Content-Disposition'] = 'attachment; filename=export.zip'
      return response 
      # for root, dirs, files in os.walk(settings.MEDIA_ROOT):
      #   for f in files:
      #     # export_zip.write(os.path.join(root, f))

    def delete_items_photoes(self, request, queryset):
      for item in queryset:
        item.images.all().delete()

    def delete_items_features(self, request, queryset):
      for item in queryset:
        item.features.all().delete()

    export_items.short_description          = "Простий експорт обраних товарів"
    export_items_to_xlsx.short_description  = "Експорт обраних товарів в ексель"
    export_items_photoes.short_description  = "Експорт у обраних товарів всіх фотографій"
    delete_items_photoes.short_description  = "Видалити у обраних товарів всі фотографій"
    delete_items_features.short_description = 'Видалити у обраних товарів всі характеристики'



