from modeltranslation.admin import (
    TranslationAdmin,
    TabbedTranslationAdmin,
    TabbedExternalJqueryTranslationAdmin,
    TabbedDjangoJqueryTranslationAdmin,
    TranslationTabularInline,
    TranslationStackedInline,
    TranslationGenericTabularInline,
    TranslationGenericStackedInline,
)
from django.conf import settings
from django.forms import TextInput, Textarea, NumberInput
from django.contrib import admin 
from django.shortcuts import reverse 
from django.utils.safestring import mark_safe
from django.contrib.admin.widgets import AdminFileWidget
from django.http import HttpResponse

from import_export.admin import ImportExportModelAdmin

from box.shop.item.models import * 
from box.shop.cart.models import * 

from box.admin import custom_admin

from datetime import timedelta, datetime
from openpyxl import Workbook


CURRENT_DOMEN = settings.CURRENT_DOMEN


class ExportCsvMixin:

    def write_items_to_xlsx(self, request, queryset):
        items = Item.objects.all()
        categories = ItemCategory.objects.all()
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        workbook.remove(workbook.active)





        worksheet1 = workbook.create_sheet(
            title='Товары',
            index=1,
        )
        columns = [
            "Мета_Заголовок",
            "Мета_Описание",
            "Мета_Ключевые_Слова",
            "Заголовок",
            "Описание",
            "Артикул",
            "Категория",
            "Ссылка",
            "Старая_Цена",
            "Новая_Цена",
            "Изображения",
        ]
        biggest_item = items.first()
        for item in items:
            if item.features.all().count() > biggest_item.features.all().count():
                biggest_item = item 
        for i in range(int(biggest_item.features.all().count())):
            columns.append("Название_Характеристики")
            columns.append("Значение_Характеристики")
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet1.cell(row=row_num, column=col_num)
            cell.value = column_title
        for item in items:
            row_num += 1
            row = [
                item.meta_title,
                item.meta_descr,
                item.meta_key,
                item.title,
                item.description,
                item.code,
                item.category.slug,
                item.slug,
                item.old_price,
                item.new_price,
                ','.join([image.image.url for image in item.images.all()]),
            ]
            features = []
            for feature in item.features.all():
                features.append(feature.name)
                features.append(feature.value)
            row.extend(features)
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet1.cell(row=row_num, column=col_num)
                cell.value = cell_value





        worksheet2 = workbook.create_sheet(
            title='Категории',
            index=2,
        )
        columns = [
            "title",
            "slug",
            "parent",
        ]
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = column_title
        for category in categories:
            row_num += 1
            parent_slug = ''
            if category.parent:
                parent_slug = category.parent.slug
            row = [
                category.title,
                category.slug,
                parent_slug,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet2.cell(row=row_num, column=col_num)
                cell.value = cell_value







        workbook.save(response)
        return response
        
    def export_items(self, request, queryset):
        meta = self.model._meta
        field_names = ['Места'] + [field.name for field in meta.fields]
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = f'attachement; filename={meta}.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in queryset:
            seats = ','.join([seat_in_order.seat.number for seat_in_order in SeatInOrder.objects.filter(order=obj)])
            # row = writer.writerow([getattr(obj, field) for field in field_names[1:]])
            writer.writerow([seats] + [getattr(obj, field) for field in field_names[1:]])
        return response
        export_items.short_description = "Експортувати вибрані товари"


class AdminImageWidget(AdminFileWidget):
  def render(self, name, value, attrs=None, renderer=None):
    output = []
    if value and getattr(value, "url", None):
      image_url = value.url
      file_name = str(value)
      output.append(
        f' <a href="{image_url}" target="_blank">'
        f'  <img src="{image_url}" alt="{file_name}" width="150" height="150" '
        f'style="object-fit: cover;"/> </a>')
    output.append(super(AdminFileWidget, self).render(name, value, attrs, renderer))
    return mark_safe(u''.join(output))


class ItemImageInline(admin.StackedInline):
    
    model = ItemImage
    extra = 0
    classes = ['collapse']
    fields = [
        'id',
        'image',
        'alt',
        'order',
    ]
    formfield_overrides = {models.ImageField: {'widget': AdminImageWidget}}


class ItemReviewInline(admin.TabularInline):
    model = ItemReview
    extra = 0 
    classes = ['collapse']
    exclude = [

    ]


class ItemInline(admin.TabularInline):
    def show_title(self, obj):
      option = "change" # "delete | history | change"
      massiv = []
      app   = obj._meta.app_label
      model = obj._meta.model_name
      url   = f'admin:{app}_{model}_{option}'
      href  = reverse(url, args=(obj.pk,))
      name  = f'{obj.title}'
      link  = mark_safe(f"<a href={href}>{name}</a>")
      return link
    show_title.short_description = 'Товар'
    model = Item 
    extra = 0
    fields = [
        'show_title',
        'new_price',
        'old_price',
        'currency',
    ]
    readonly_fields = [
        'show_title',
        'new_price',
        'old_price',
        'currency',
    ]
    classes = ['collapse']
    if settings.MULTIPLE_CATEGORY:
        filter_horizontal = [
            'categories',
        ]
    else:
        filter_horizontal = [
            'category',
        ]




class ItemCategoryInline(admin.TabularInline):
    model = ItemCategory 
    extra = 0
    exclude = [
        'meta_title',
        'meta_descr',
        'meta_key',
        'description',
        'code',
        'created',
        'updated',
    ]
    classes = ['collapse']
    verbose_name = "підкатегорія"
    verbose_name_plural = "підкатегорії"
    prepopulated_fields = {
        "slug": ("title",), 
    }


class ItemFeatureInline(admin.StackedInline):
    model = ItemFeature
    # model = ItemFeature.items.through
    extra = 0
    classes = ['collapse']
    exclude = [
        'code',
        # 'category',
        'categories',
    ]



def get_fieldsets():
    if settings.MULTIPLE_CATEGORY:
        fieldsets = (
            ('ТОВАР', {
                'fields':(
                    (
                    'in_stock',
                    # 'is_new',
                    'is_active',
                    ),
                    (
                    'title',
                    'code',
                    ),
                    (
                    'old_price',
                    'new_price',
                    'currency',
                    ),
                    'description',
                    'thumbnail',
                    'categories',
                    'created',
                    'updated',
                ),
            }),
            ('SEO', {
                'fields':(
                    'slug',
                    (
                    'meta_title',
                    'meta_descr',
                    'meta_key',
                    ),
                ),
                'classes':(
                    'collapse', 
                    'wide',
                    'extrapretty',
                ),
            }),
        )
    else:
        fieldsets = (
            ('ТОВАР', {
                'fields':(
                    (
                    'in_stock',
                    # 'is_new',
                    'is_active',
                    ),
                    (
                    'title',
                    'code',
                    ),
                    (
                    'old_price',
                    'new_price',
                    'currency',
                    ),
                    'description',
                    'thumbnail',
                    'category',
                    'created',
                    'updated',
                ),
            }),
            ('SEO', {
                'fields':(
                    'slug',
                    (
                    'meta_title',
                    'meta_descr',
                    'meta_key',
                    ),
                ),
                'classes':(
                    'collapse', 
                    'wide',
                    'extrapretty',
                ),
            }),
        )
    
    return fieldsets


@admin.register(Item, site=custom_admin)
class ItemAdmin(admin.ModelAdmin, ExportCsvMixin):
    # class Media:
    #     js = (
    #         'http://ajax.googleapis.com/ajax/libs/jquery/1.9.1/jquery.min.js',
    #         'http://ajax.googleapis.com/ajax/libs/jqueryui/1.10.2/jquery-ui.min.js',
    #         'modeltranslation/js/tabbed_translation_fields.js',
    #     )
    #     css = {
    #         'screen': ('modeltranslation/css/tabbed_translation_fields.css',),
    #     }
    actions = [
        'export_items',
        'write_items_to_xlsx'
    ]
    change_list_template = 'items_change_list.html'
    # TODO: static method 
    fieldsets = get_fieldsets()

    formfield_overrides = {
        models.CharField: {'widget': NumberInput(attrs={'size':'20'})},
        models.CharField: {'widget': TextInput(attrs={'size':'20'})},
        models.TextField: {'widget': Textarea(attrs={'rows':6, 'cols':20})},
    }
    prepopulated_fields = {
        "slug": ("title",),
        "code": ("title",),
    }
    exclude = []
    readonly_fields = [
        'created',
        'updated',
    ]
    save_on_top = True 
    save_on_bottom = True 
    search_fields = [
        'title',
        'code',
        'slug',
        'description',
    ]
    list_filter = [
        "in_stock",
        "is_new", 
        "is_active", 
        # "category",
    ]
    list_display = [
        'id',
        'title',
        'category',
        'price',
        'old_price',
        'in_stock',
        # 'is_new',
        'is_active',
    ]
    list_editable = [
        'category',
    ]
    list_display_links = [
        'id', 
        'title',
    ]
    inlines = [
        ItemImageInline,
        ItemFeatureInline,
        ItemReviewInline, 
    ]
    list_per_page = 20


@admin.register(ItemCategory, site=custom_admin)
class ItemCategoryAdmin(admin.ModelAdmin):
    inlines = [
        # ItemInline,
        ItemCategoryInline,
    ]
    list_display = [
        'id',
        'tree_title',
        'slug',
        'currency',
    ]
    list_display_links = [
        'id',
        'tree_title',
    ]
    list_editable= [
        'currency',
        'slug'
    ]
    fieldsets = (
        ('ОСНОВНА ІНФОРМАЦІЯ', {
            "fields":(
                "title",
                "thumbnail",
                "is_active",
                "created",
                'updated',
                'parent',
                "currency",
            ),
            'classes':(
                'wide',
            )
        }),
        ("SEO",{
            "fields":(
                (
                'meta_title',
                'meta_descr',
                'meta_key',
                ),
                "slug",
            ),
            'classes':(
                'collapse',
                'wide'
            )
        }),
    )
    readonly_fields = [
        'created',
        'updated',
    ]
    formfield_overrides = {
        models.CharField: {'widget': NumberInput(attrs={'size':'40'})},
        models.CharField: {'widget': TextInput(attrs={'size':'40'})},
        models.TextField: {'widget': Textarea(attrs={'rows':6, 'cols':20})},
    }
    prepopulated_fields = {
        "slug": ("title",),
    }


# @admin.register(ItemImage, site=custom_admin)
class ItemImageAdmin(admin.ModelAdmin):
    save_on_top = True 
    save_on_bottom = True 
    def headshot_image(self, obj):
        return mark_safe(
            f'<img \
                src="{obj.headshot.url}" \
                width="{obj.headshot.width}" \
                height={obj.headshot.height} \
            />'
        )
    def show_item(self, obj):
      option = "change" # "delete | history | change"
      massiv = []
      obj   = obj.item
      app   = obj._meta.app_label
      model = obj._meta.model_name
      url   = f'admin:{app}_{model}_{option}'
      href  = reverse(url, args=(obj.pk,))
      name  = f'{obj.title}'
      link  = mark_safe(f"<a href={href}>{name}</a>")
      return link
    def show_image(self, obj):
      option = "change" # "delete | history | change"
      massiv = []
      app   = obj._meta.app_label
      model = obj._meta.model_name
      url   = f'admin:{app}_{model}_{option}'
      href  = reverse(url, args=(obj.pk,))
      name  = f'{obj.image}'
      link  = mark_safe(f"<a href={href}>{name}</a>")
      return link
    show_item.short_description = ('Товар')
    list_display = [
        'id',
        'show_image',
        'alt',
        'show_item',
    ]
    list_display_links = [
        'id',
        'alt',
    ]


# @admin.register(ItemReview, site=custom_admin)
class ItemReviewAadmin(admin.ModelAdmin):
    pass


# @admin.register(Currency, site=custom_admin)
class CurrencyAdmin(admin.ModelAdmin):
    list_display_links = [
        'id',
        'name',
    ]
    list_display = [
        'id',
        'name',
        'is_main',
    ]


       
@admin.register(CurrencyRatio, site=custom_admin)
class CurrencyRatioAdmin(admin.ModelAdmin):
    list_display_links = [
        'id',
        'main',
        'compared',
        'ratio',
    ]
    list_display = [
        'id',
        'main',
        'compared',
        'ratio',
    ]

